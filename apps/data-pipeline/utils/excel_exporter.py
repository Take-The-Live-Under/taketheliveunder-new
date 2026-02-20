"""
Excel Exporter for Referee Analysis
Generates comprehensive Excel workbooks with charts and visualizations
"""

import pandas as pd
import logging
from datetime import datetime
from typing import Dict
from openpyxl import Workbook
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports referee analysis data to Excel with charts and visualizations"""

    def __init__(self, output_path: str):
        """
        Initialize exporter

        Args:
            output_path: Path to output Excel file
        """
        self.output_path = output_path

    def export(self, data: Dict) -> str:
        """
        Export referee analysis to Excel workbook

        Args:
            data: Dictionary containing:
                - referee_stats: DataFrame with referee statistics
                - game_logs: DataFrame with game logs
                - summary: Summary statistics dict

        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating Excel report: {self.output_path}")

        referee_stats = data["referee_stats"]
        game_logs = data["game_logs"]
        summary = data["summary"]

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        self._create_summary_sheet(wb, referee_stats, summary)
        self._create_raw_data_sheet(wb, game_logs)
        self._create_referee_profiles_sheet(wb, referee_stats)
        self._create_charts_sheet(wb, referee_stats, game_logs)

        # Save workbook
        wb.save(self.output_path)
        logger.info(f"Excel report saved to: {self.output_path}")

        return self.output_path

    def _create_summary_sheet(self, wb: Workbook, referee_stats: pd.DataFrame, summary: Dict):
        """Create summary dashboard sheet"""
        ws = wb.create_sheet("Summary Dashboard", 0)

        # Header
        ws['A1'] = "NCAA Basketball Referee Analysis - Summary Dashboard"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Overall stats
        row = 4
        ws[f'A{row}'] = "Overall Statistics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        for key, value in summary.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1

        # Top 10 referees by avg fouls/game
        row += 2
        ws[f'A{row}'] = "Top 10 Referees by Average Fouls/Game"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        top_10 = referee_stats.head(10)
        for idx, (_, ref) in enumerate(top_10.iterrows(), start=row):
            ws[f'A{ref + row - row}'] = ref['referee']
            ws[f'B{ref + row - row}'] = ref['avg_total_fouls']
            ws[f'C{ref + row - row}'] = ref['total_games']

        # Column headers for top 10
        ws[f'A{row}'] = "Referee"
        ws[f'B{row}'] = "Avg Fouls/Game"
        ws[f'C{row}'] = "Total Games"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'].font = Font(bold=True)

        # Top 10 data
        for idx, (_, ref_row) in enumerate(top_10.iterrows(), start=row + 1):
            ws[f'A{idx}'] = ref_row['referee']
            ws[f'B{idx}'] = ref_row['avg_total_fouls']
            ws[f'C{idx}'] = ref_row['total_games']

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

    def _create_raw_data_sheet(self, wb: Workbook, game_logs: pd.DataFrame):
        """Create raw data table sheet"""
        ws = wb.create_sheet("Raw Data")

        # Add header
        ws['A1'] = "Complete Game Logs - All Referee Assignments"
        ws['A1'].font = Font(size=14, bold=True)

        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(game_logs, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Format header row
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze panes
        ws.freeze_panes = 'A4'

    def _create_referee_profiles_sheet(self, wb: Workbook, referee_stats: pd.DataFrame):
        """Create referee profiles sheet"""
        ws = wb.create_sheet("Referee Profiles")

        # Add header
        ws['A1'] = "Referee Statistical Profiles"
        ws['A1'].font = Font(size=14, bold=True)

        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(referee_stats, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Format header row
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze panes
        ws.freeze_panes = 'A4'

    def _create_charts_sheet(self, wb: Workbook, referee_stats: pd.DataFrame, game_logs: pd.DataFrame):
        """Create charts and visualizations sheet"""
        ws = wb.create_sheet("Charts & Visualizations")

        # Header
        ws['A1'] = "Referee Analysis Visualizations"
        ws['A1'].font = Font(size=14, bold=True)

        # Chart 1: Top 15 Referees by Avg Fouls/Game (Bar Chart)
        self._create_top_referees_bar_chart(ws, referee_stats)

        # Chart 2: Home vs Away Foul Averages (Scatter Chart)
        self._create_home_away_scatter_chart(ws, referee_stats)

        # Add data tables for charts at the bottom
        self._add_chart_data_tables(ws, referee_stats)

    def _create_top_referees_bar_chart(self, ws, referee_stats: pd.DataFrame):
        """Create bar chart of top referees by avg fouls"""
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.title = "Top 15 Referees by Average Fouls Per Game"
        chart.x_axis.title = "Referee"
        chart.y_axis.title = "Average Fouls Per Game"

        # Get top 15
        top_15 = referee_stats.head(15)

        # Add data to worksheet (starting at row 20 to leave space for charts)
        start_row = 20
        ws[f'A{start_row}'] = "Referee"
        ws[f'B{start_row}'] = "Avg Fouls/Game"

        for idx, (_, row) in enumerate(top_15.iterrows(), start=start_row + 1):
            ws[f'A{idx}'] = row['referee']
            ws[f'B{idx}'] = row['avg_total_fouls']

        # Create chart
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(top_15))
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(top_15))

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Style
        chart.width = 20
        chart.height = 12

        # Add to worksheet
        ws.add_chart(chart, "D3")

    def _create_home_away_scatter_chart(self, ws, referee_stats: pd.DataFrame):
        """Create scatter plot of home vs away foul averages"""
        # Create scatter chart
        chart = ScatterChart()
        chart.title = "Home vs Away Foul Averages by Referee"
        chart.x_axis.title = "Average Home Fouls"
        chart.y_axis.title = "Average Away Fouls"

        # Add data to worksheet (starting at row 40)
        start_row = 40
        ws[f'D{start_row}'] = "Avg Home Fouls"
        ws[f'E{start_row}'] = "Avg Away Fouls"

        for idx, (_, row) in enumerate(referee_stats.iterrows(), start=start_row + 1):
            ws[f'D{idx}'] = row['avg_home_fouls']
            ws[f'E{idx}'] = row['avg_away_fouls']

        # Create chart
        xvalues = Reference(ws, min_col=4, min_row=start_row + 1, max_row=start_row + len(referee_stats))
        yvalues = Reference(ws, min_col=5, min_row=start_row + 1, max_row=start_row + len(referee_stats))

        series = Series(yvalues, xvalues)
        chart.series.append(series)

        # Style
        chart.width = 15
        chart.height = 12

        # Add to worksheet
        ws.add_chart(chart, "D22")

    def _add_chart_data_tables(self, ws, referee_stats: pd.DataFrame):
        """Add data tables used for charts to the bottom of the sheet"""
        # This data is already added in the chart creation methods
        # Just add labels
        ws['A19'] = "Chart Data: Top 15 Referees"
        ws['A19'].font = Font(bold=True, size=12)

        ws['D39'] = "Chart Data: Home vs Away Comparison (All Referees)"
        ws['D39'].font = Font(bold=True, size=12)


def get_excel_exporter(output_path: str) -> ExcelExporter:
    """
    Get instance of ExcelExporter

    Args:
        output_path: Path to output Excel file

    Returns:
        ExcelExporter instance
    """
    return ExcelExporter(output_path)
