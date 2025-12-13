#!/usr/bin/env python3
"""
Predict Tomorrow's NCAA Basketball Game Totals

Fetches upcoming games, loads trained models, and predicts total points
with individual team score breakdowns.

Usage:
    python predict_totals.py

Outputs:
    - Excel file: data/predictions/predictions_YYYY-MM-DD.xlsx
    - CSV file: data/predictions/predictions_YYYY-MM-DD.csv
"""

import pandas as pd
import numpy as np
import pickle
import requests
import os
from pathlib import Path
from datetime import datetime, timedelta
import sys

# Add models to path
sys.path.insert(0, str(Path(__file__).parent))

from models.data_pipeline.historical_games_processor import normalize_team_name


# ============================================================================
# CONFIGURATION
# ============================================================================

DATA_DIR = Path(__file__).parent / "data"
PREDICTIONS_DIR = DATA_DIR / "predictions"
PREDICTIONS_DIR.mkdir(parents=True, exist_ok=True)

MODEL_PATH = DATA_DIR / "trained_models" / "ensemble_latest.pkl"
KENPOM_PATH = DATA_DIR / "kenpom_historical" / "season_2025" / "cleaned_kenpom_data_latest.csv"

# The Odds API
ODDS_API_KEY = os.getenv('ODDS_API_KEY', 'c1e957e22dfde2c23b3cac82758bef3e')
ODDS_API_BASE = "https://api.the-odds-api.com/v4"

# Sport mode - check environment variable first
SPORT_MODE = os.getenv('SPORT_MODE', 'basketball_ncaab')


# ============================================================================
# FETCH UPCOMING GAMES
# ============================================================================

def fetch_upcoming_games(date_str: str = None):
    """
    Fetch upcoming NCAA basketball games.

    Args:
        date_str: Date to fetch games for (YYYY-MM-DD). Defaults to tomorrow.

    Returns:
        List of game dicts with home_team, away_team, commence_time
    """
    if date_str is None:
        # Default to tomorrow
        tomorrow = datetime.now() + timedelta(days=1)
        date_str = tomorrow.strftime("%Y-%m-%d")

    print(f"\nFetching games for {date_str}...")

    url = f"{ODDS_API_BASE}/sports/{SPORT_MODE}/odds"
    params = {
        'apiKey': ODDS_API_KEY,
        'regions': 'us',
        'markets': 'totals',
        'oddsFormat': 'american'
    }

    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        games_data = response.json()

        print(f"  Found {len(games_data)} upcoming games")

        # Filter to games on the target date
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        games = []

        for game in games_data:
            game_time = datetime.fromisoformat(game['commence_time'].replace('Z', '+00:00'))
            game_date = game_time.date()

            if game_date == target_date:
                home_team = game['home_team']
                away_team = game['away_team']

                # Extract over/under line if available
                ou_line = None
                if 'bookmakers' in game and game['bookmakers']:
                    for bookmaker in game['bookmakers']:
                        if 'markets' in bookmaker:
                            for market in bookmaker['markets']:
                                if market['key'] == 'totals' and 'outcomes' in market:
                                    for outcome in market['outcomes']:
                                        if 'point' in outcome:
                                            ou_line = outcome['point']
                                            break
                                if ou_line:
                                    break
                        if ou_line:
                            break

                games.append({
                    'home_team': home_team,
                    'away_team': away_team,
                    'commence_time': game_time.strftime("%Y-%m-%d %H:%M"),
                    'ou_line': ou_line
                })

        print(f"  Games on {date_str}: {len(games)}")
        return games

    except Exception as e:
        print(f"  ERROR fetching games: {e}")
        return []


# ============================================================================
# LOAD KENPOM DATA
# ============================================================================

def load_kenpom_ratings():
    """Load KenPom season ratings"""
    print(f"\nLoading KenPom ratings from: {KENPOM_PATH}")

    if not KENPOM_PATH.exists():
        raise FileNotFoundError(f"KenPom data not found at {KENPOM_PATH}")

    df = pd.read_csv(KENPOM_PATH)
    print(f"  Loaded {len(df)} teams")

    # Normalize team names
    df['team_normalized'] = df['team'].apply(normalize_team_name)

    return df


def find_team_in_kenpom(team_name: str, kenpom_df: pd.DataFrame):
    """Find team in KenPom data with flexible matching"""
    team_normalized = normalize_team_name(team_name)

    # Try exact match
    exact_match = kenpom_df[kenpom_df['team_normalized'] == team_normalized]
    if not exact_match.empty:
        return exact_match.iloc[0]

    # Try partial match
    for idx, row in kenpom_df.iterrows():
        kp_name = row['team_normalized'].lower()
        tm_name = team_normalized.lower()

        if kp_name in tm_name or tm_name in kp_name:
            return row

    return None


# ============================================================================
# MAKE PREDICTIONS
# ============================================================================

def load_model():
    """Load trained ensemble model"""
    print(f"\nLoading trained model from: {MODEL_PATH}")

    if not MODEL_PATH.exists():
        raise FileNotFoundError(f"Trained model not found at {MODEL_PATH}")

    with open(MODEL_PATH, 'rb') as f:
        ensemble = pickle.load(f)

    print(f"  Model loaded: {ensemble.name}")
    return ensemble


def prepare_game_features(home_team, away_team, kenpom_df):
    """
    Prepare features for a single game.

    Returns:
        DataFrame with features, or None if teams not found
    """
    # Find teams in KenPom data
    home_kenpom = find_team_in_kenpom(home_team, kenpom_df)
    away_kenpom = find_team_in_kenpom(away_team, kenpom_df)

    if home_kenpom is None or away_kenpom is None:
        return None

    # Build feature dict
    features = {}

    # Add team 1 (home) features
    for col in kenpom_df.columns:
        if col not in ['team', 'team_normalized']:
            features[f'team_1_{col}'] = home_kenpom[col]

    # Add team 2 (away) features
    for col in kenpom_df.columns:
        if col not in ['team', 'team_normalized']:
            features[f'team_2_{col}'] = away_kenpom[col]

    # Add home indicator
    features['team_1_is_home'] = True

    return pd.DataFrame([features])


def predict_games(games, ensemble, kenpom_df):
    """
    Predict total points for all games.

    Returns:
        DataFrame with predictions and breakdowns
    """
    print(f"\nMaking predictions for {len(games)} games...")

    predictions = []

    for game in games:
        home_team = game['home_team']
        away_team = game['away_team']

        # Prepare features
        features = prepare_game_features(home_team, away_team, kenpom_df)

        if features is None:
            print(f"  ⚠️  Skipping {home_team} vs {away_team} (teams not found in KenPom)")
            continue

        # Use Pomeroy predictor directly (more reliable than ensemble for small training set)
        pomeroy_details = ensemble.pomeroy.predict_with_details(features)

        # Calculate confidence based on tempo and efficiency certainty
        # Higher confidence for teams with stable metrics
        home_adjem = features['team_1_adjem'].values[0] if 'team_1_adjem' in features else 0
        away_adjem = features['team_2_adjem'].values[0] if 'team_2_adjem' in features else 0
        avg_strength = (abs(home_adjem) + abs(away_adjem)) / 2
        confidence = min(100, 50 + avg_strength * 2)  # Higher for better teams

        prediction = {
            'date': game['commence_time'],
            'home_team': home_team,
            'away_team': away_team,
            'home_projected_score': round(pomeroy_details['team1_points'].values[0], 1),
            'away_projected_score': round(pomeroy_details['team2_points'].values[0], 1),
            'projected_total': round(pomeroy_details['total_points'].values[0], 1),
            'home_efficiency': round(pomeroy_details['team1_efficiency'].values[0], 1),
            'away_efficiency': round(pomeroy_details['team2_efficiency'].values[0], 1),
            'confidence': round(confidence, 0),
            'ou_line': game.get('ou_line'),
            'projected_tempo': round(pomeroy_details['possessions'].values[0], 1),
        }

        # Add over/under comparison if line is available
        if prediction['ou_line'] is not None:
            diff = prediction['projected_total'] - prediction['ou_line']
            prediction['vs_line'] = round(diff, 1)
            if diff > 2:
                prediction['suggestion'] = 'OVER'
            elif diff < -2:
                prediction['suggestion'] = 'UNDER'
            else:
                prediction['suggestion'] = 'PASS'
        else:
            prediction['vs_line'] = None
            prediction['suggestion'] = 'NO LINE'

        predictions.append(prediction)

    df = pd.DataFrame(predictions)
    print(f"  Successfully predicted {len(df)} games")

    return df


# ============================================================================
# EXPORT TO EXCEL
# ============================================================================

def export_to_excel(predictions_df, date_str):
    """Export predictions to Excel with nice formatting"""
    print(f"\nExporting to Excel...")

    # Create filename
    excel_path = PREDICTIONS_DIR / f"predictions_{date_str}.xlsx"
    csv_path = PREDICTIONS_DIR / f"predictions_{date_str}.csv"

    # Save CSV
    predictions_df.to_csv(csv_path, index=False)
    print(f"  Saved CSV: {csv_path}")

    # Save Excel with formatting
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        ws = wb.active
        ws.title = f"Predictions {date_str}"

        # Add title
        ws['A1'] = f"NCAA Basketball Predictions - {date_str}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:K1')

        # Add header row
        headers = [
            'Game Time', 'Home Team', 'Away Team',
            'Home Score', 'Away Score', 'Total',
            'O/U Line', 'vs Line', 'Suggestion',
            'Confidence', 'Tempo', 'Home Eff', 'Away Eff'
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Add data rows
        for row_idx, row in enumerate(predictions_df.itertuples(), 4):
            ws.cell(row=row_idx, column=1, value=row.date)
            ws.cell(row=row_idx, column=2, value=row.home_team)
            ws.cell(row=row_idx, column=3, value=row.away_team)
            ws.cell(row=row_idx, column=4, value=row.home_projected_score)
            ws.cell(row=row_idx, column=5, value=row.away_projected_score)
            ws.cell(row=row_idx, column=6, value=row.projected_total)
            ws.cell(row=row_idx, column=7, value=row.ou_line if not pd.isna(row.ou_line) else 'N/A')
            ws.cell(row=row_idx, column=8, value=row.vs_line if not pd.isna(row.vs_line) else 'N/A')

            # Color code suggestion
            suggestion_cell = ws.cell(row=row_idx, column=9, value=row.suggestion)
            if row.suggestion == 'OVER':
                suggestion_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif row.suggestion == 'UNDER':
                suggestion_cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')

            ws.cell(row=row_idx, column=10, value=row.confidence)
            ws.cell(row=row_idx, column=11, value=row.projected_tempo)
            ws.cell(row=row_idx, column=12, value=row.home_efficiency)
            ws.cell(row=row_idx, column=13, value=row.away_efficiency)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(excel_path)
        print(f"  Saved Excel: {excel_path}")

    except Exception as e:
        print(f"  Warning: Could not create Excel with formatting: {e}")
        # Fallback to simple Excel
        predictions_df.to_excel(excel_path, index=False, engine='openpyxl')
        print(f"  Saved basic Excel: {excel_path}")

    return excel_path, csv_path


# ============================================================================
# MAIN
# ============================================================================

def main():
    """Main prediction pipeline"""
    print("="*80)
    print("NCAA Basketball Game Total Predictions")
    print("="*80)

    # Get target date (tomorrow by default)
    tomorrow = datetime.now() + timedelta(days=1)
    date_str = tomorrow.strftime("%Y-%m-%d")

    # Fetch upcoming games
    games = fetch_upcoming_games(date_str)

    if not games:
        print(f"\nNo games found for {date_str}")
        return False

    # Load model and data
    ensemble = load_model()
    kenpom_df = load_kenpom_ratings()

    # Make predictions
    predictions_df = predict_games(games, ensemble, kenpom_df)

    if predictions_df.empty:
        print("\nNo valid predictions generated")
        return False

    # Export results
    excel_path, csv_path = export_to_excel(predictions_df, date_str)

    # Print summary
    print("\n" + "="*80)
    print("Prediction Summary")
    print("="*80)
    print(f"Games predicted: {len(predictions_df)}")
    print(f"Average projected total: {predictions_df['projected_total'].mean():.1f}")
    print(f"Average confidence: {predictions_df['confidence'].mean():.0f}/100")

    if predictions_df['ou_line'].notna().any():
        over_picks = (predictions_df['suggestion'] == 'OVER').sum()
        under_picks = (predictions_df['suggestion'] == 'UNDER').sum()
        print(f"\nSuggestions: {over_picks} OVER, {under_picks} UNDER")

    print(f"\n✅ Results saved to:")
    print(f"   Excel: {excel_path}")
    print(f"   CSV: {csv_path}")

    # Display predictions table
    print("\n" + "="*80)
    print("Predictions Preview")
    print("="*80)
    display_cols = ['home_team', 'away_team', 'home_projected_score',
                    'away_projected_score', 'projected_total', 'ou_line', 'suggestion']
    print(predictions_df[display_cols].to_string(index=False))

    return True


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
